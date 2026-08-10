"""销售订单导入会话：规则解析 + 人工确认，不清不猜。"""

from pathlib import Path

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.config import get_settings
from app.db import Base
from app.models import OwnProduct, Partner, Tenant
from app.services.sales_order_ai_import import (
    apply_clarifications,
    confirm_import_session,
    create_import_session,
    patch_import_draft,
)
from app.services.sales_order_import import build_sales_order_import_template_bytes
from app.services.sales_order_service import SalesOrderError


@pytest.fixture()
def db(tmp_path, monkeypatch):
    uploads = tmp_path / "uploads"
    uploads.mkdir()
    monkeypatch.setenv("UPLOADS_DIR", str(uploads))
    get_settings.cache_clear()

    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()
    tenant = Tenant(name="导入厂")
    session.add(tenant)
    session.flush()
    session.commit()
    yield session, tenant.id
    session.close()
    get_settings.cache_clear()


def _seed_products(session, tenant_id: int, codes: list[str]) -> dict[str, int]:
    out: dict[str, int] = {}
    for code in codes:
        p = OwnProduct(tenant_id=tenant_id, product_code=code, is_active=True)
        session.add(p)
        session.flush()
        out[code] = p.id
    session.commit()
    return out


def test_import_session_needs_customer_when_ambiguous(db):
    session, tenant_id = db
    session.add(
        Partner(
            tenant_id=tenant_id,
            name="红丫丫鞋业总店",
            short_name="红丫丫鞋业",
            is_customer=True,
            is_active=True,
        )
    )
    session.add(
        Partner(
            tenant_id=tenant_id,
            name="红丫丫鞋业分厂",
            short_name="红丫丫鞋业",
            is_customer=True,
            is_active=True,
        )
    )
    _seed_products(session, tenant_id, ["21L533-25", "21L527-20"])

    path = Path(__file__).resolve().parents[1] / "assets" / "templates" / "sales_order_import.xlsx"
    pub = create_import_session(
        session,
        tenant_id,
        filename="订单.xlsx",
        content=path.read_bytes(),
        created_by=None,
    )
    assert pub["status"] == "needs_clarification"
    assert not pub["can_confirm"]
    types = {q["type"] for q in pub["clarifications"]}
    assert "customer" in types


def test_import_session_ready_and_confirm(db):
    session, tenant_id = db
    cust = Partner(
        tenant_id=tenant_id,
        name="示例客户有限公司",
        short_name="示例客户",
        is_customer=True,
        is_active=True,
    )
    session.add(cust)
    _seed_products(session, tenant_id, ["OP-001"])

    content = build_sales_order_import_template_bytes()
    pub = create_import_session(
        session,
        tenant_id,
        filename="模版.xlsx",
        content=content,
        created_by=None,
    )
    assert pub["draft"]["customer"]["status"] == "matched"
    assert pub["draft"]["customer"]["customer_id"] == cust.id
    assert all(ln["product_status"] == "matched" for ln in pub["draft"]["lines"])
    assert pub["can_confirm"] is True
    assert pub["status"] == "ready_for_review"

    confirmed = confirm_import_session(session, tenant_id, pub["id"], created_by=None)
    assert confirmed["status"] == "confirmed"
    assert confirmed["result"]["order_no"]
    assert confirmed["result"]["sales_order_id"]


def test_import_session_clarify_product_then_confirm(db):
    session, tenant_id = db
    cust = Partner(
        tenant_id=tenant_id,
        name="示例客户",
        is_customer=True,
        is_active=True,
    )
    session.add(cust)
    # 故意不建 OP-001，只建别的产品，强制产品澄清
    other = OwnProduct(tenant_id=tenant_id, product_code="OTHER-9", is_active=True)
    session.add(other)
    session.commit()

    content = build_sales_order_import_template_bytes()
    pub = create_import_session(
        session,
        tenant_id,
        filename="模版.xlsx",
        content=content,
        created_by=None,
    )
    assert pub["status"] == "needs_clarification"
    product_qs = [q for q in pub["clarifications"] if q["type"] == "product"]
    assert product_qs

    answers = [
        {"id": q["id"], "own_product_id": other.id} for q in product_qs
    ]
    pub2 = apply_clarifications(session, tenant_id, pub["id"], answers)
    assert pub2["can_confirm"] is True

    confirmed = confirm_import_session(session, tenant_id, pub2["id"], created_by=None)
    assert confirmed["status"] == "confirmed"


def test_import_parse_fail_without_ai_key(db, monkeypatch):
    session, tenant_id = db
    monkeypatch.setattr(
        "app.services.sales_order_ai_import._ai_available",
        lambda: False,
    )
    # 空表 / 无表头
    from openpyxl import Workbook
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "随便"
    ws["B1"] = "几列"
    buf = BytesIO()
    wb.save(buf)
    pub = create_import_session(
        session,
        tenant_id,
        filename="bad.xlsx",
        content=buf.getvalue(),
        created_by=None,
    )
    assert pub["status"] == "failed"
    assert pub["clarifications"]
    assert pub["clarifications"][0]["type"] == "parse"
    assert not pub["can_confirm"]


def test_patch_customer_name_without_override_field(db):
    """客户只能在核对草稿里改，不再有上传时覆盖参数。"""
    session, tenant_id = db
    session.add(
        Partner(
            tenant_id=tenant_id,
            name="示例客户",
            is_customer=True,
            is_active=True,
        )
    )
    _seed_products(session, tenant_id, ["OP-001"])
    pub = create_import_session(
        session,
        tenant_id,
        filename="模版.xlsx",
        content=build_sales_order_import_template_bytes(),
        created_by=None,
    )
    patched = patch_import_draft(
        session,
        tenant_id,
        pub["id"],
        {"customer": {"customer_id": None, "customer_name": "手填客户甲"}},
    )
    assert patched["draft"]["customer"]["customer_name"] == "手填客户甲"
    assert patched["draft"]["customer"]["status"] == "matched"
    assert patched["can_confirm"] is True


def test_confirm_blocked_when_needs_clarification(db):
    session, tenant_id = db
    _seed_products(session, tenant_id, ["OP-001"])
    # 无客户档案 → needs_input
    pub = create_import_session(
        session,
        tenant_id,
        filename="模版.xlsx",
        content=build_sales_order_import_template_bytes(),
        created_by=None,
    )
    assert pub["status"] == "needs_clarification"
    with pytest.raises(SalesOrderError) as ei:
        confirm_import_session(session, tenant_id, pub["id"], created_by=None)
    assert ei.value.code == "needs_clarification"


def test_attr_checks_flag_fabric_lining_color_mismatch(db):
    from app.models import Color, OwnProductColor

    session, tenant_id = db
    session.add(
        Partner(
            tenant_id=tenant_id,
            name="示例客户",
            is_customer=True,
            is_active=True,
        )
    )
    color_ok = Color(tenant_id=tenant_id, name="黑色", code="BK")
    session.add(color_ok)
    session.flush()
    product = OwnProduct(
        tenant_id=tenant_id,
        product_code="OP-001",
        fabric="牛皮",
        lining="猪皮",
        is_active=True,
    )
    session.add(product)
    session.flush()
    session.add(
        OwnProductColor(tenant_id=tenant_id, own_product_id=product.id, color_id=color_ok.id)
    )
    session.commit()

    pub = create_import_session(
        session,
        tenant_id,
        filename="模版.xlsx",
        content=build_sales_order_import_template_bytes(),
        created_by=None,
    )
    # 模版第1行：黑色/牛皮/猪皮 → 一致
    row0 = pub["draft"]["lines"][0]
    assert row0["attr_checks"]["color"]["status"] == "ok"
    assert row0["attr_checks"]["fabric"]["status"] == "ok"
    assert row0["attr_checks"]["lining"]["status"] == "ok"
    assert row0["has_attr_mismatch"] is False
    # 第2行米色不在产品可选色
    row1 = pub["draft"]["lines"][1]
    assert row1["attr_checks"]["color"]["status"] == "mismatch"
    assert row1["attr_checks"]["color"]["excel"] == "米色"
    assert "黑色" in (row1["attr_checks"]["color"]["system"] or "")
    assert row1["has_attr_mismatch"] is True

    patched = patch_import_draft(
        session,
        tenant_id,
        pub["id"],
        {
            "lines": [
                {
                    "index": 0,
                    "color_name": "米色",
                    "fabric": "开边珠",
                    "lining": "猪皮",
                }
            ]
        },
    )
    row = patched["draft"]["lines"][0]
    assert row["has_attr_mismatch"] is True
    assert row["attr_checks"]["color"]["status"] == "mismatch"
    assert row["attr_checks"]["fabric"]["status"] == "mismatch"
    assert row["attr_checks"]["fabric"]["excel"] == "开边珠"
    assert row["attr_checks"]["fabric"]["system"] == "牛皮"
    assert row["attr_checks"]["lining"]["status"] == "ok"
