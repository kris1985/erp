"""销售订单智能导入会话：规则解析优先 → 可选 AI 表头映射 → 人工核对确认后建单。

不清不猜：客户/产品有歧义时进入澄清，禁止静默选定。
一期图片：仅抽取浮动图，用途由用户在核对页指定（Logo / 做货要求图 / 不用）。
"""

from __future__ import annotations

import json
import re
import uuid
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.config import get_settings
from app.models import Color, OwnProduct, Partner, SalesOrder, Size
from app.schemas.api import SalesOrderCreate, SalesOrderLineIn, SalesOrderLineItemIn
from app.services.sales_order_import import (
    _ensure_product_color,
    _get_or_create_color,
    _get_or_create_size,
    parse_sales_order_workbook,
)
from app.services.sales_order_service import SalesOrderError, create_sales_order, serialize_sales_order

SESSION_TTL_HOURS = 24
STATUS_NEEDS_CLARIFICATION = "needs_clarification"
STATUS_READY = "ready_for_review"
STATUS_CONFIRMED = "confirmed"
STATUS_FAILED = "failed"

IMAGE_ROLE_IGNORE = "ignore"
IMAGE_ROLE_LOGO = "brand_logo"
IMAGE_ROLE_NOTES = "notes_image"


def _sessions_root() -> Path:
    root = Path(get_settings().uploads_dir).resolve().parent / "sales_order_import"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _tenant_dir(tenant_id: int) -> Path:
    p = _sessions_root() / str(tenant_id)
    p.mkdir(parents=True, exist_ok=True)
    return p


def _session_path(tenant_id: int, session_id: str) -> Path:
    return _tenant_dir(tenant_id) / f"{session_id}.json"


def _uploads_dir() -> Path:
    p = Path(get_settings().uploads_dir)
    p.mkdir(parents=True, exist_ok=True)
    return p


def _json_default(obj: Any) -> Any:
    if isinstance(obj, (date, datetime)):
        return obj.isoformat()
    if isinstance(obj, Decimal):
        return str(obj)
    raise TypeError(f"not serializable: {type(obj)}")


def _save_session(tenant_id: int, session: dict) -> None:
    path = _session_path(tenant_id, session["id"])
    path.write_text(
        json.dumps(session, ensure_ascii=False, indent=2, default=_json_default),
        encoding="utf-8",
    )


def _load_session(tenant_id: int, session_id: str) -> dict:
    path = _session_path(tenant_id, session_id)
    if not path.is_file():
        raise SalesOrderError("import_session_not_found", "导入会话不存在或已过期")
    data = json.loads(path.read_text(encoding="utf-8"))
    if int(data.get("tenant_id") or 0) != tenant_id:
        raise SalesOrderError("import_session_not_found", "导入会话不存在或已过期")
    created = datetime.fromisoformat(data["created_at"])
    if datetime.utcnow() - created > timedelta(hours=SESSION_TTL_HOURS):
        raise SalesOrderError("import_session_expired", "导入会话已过期，请重新上传")
    return data


def _norm_name(s: str) -> str:
    text = (s or "").strip().lower()
    text = re.sub(r"\s+", "", text)
    for suffix in ("订单明细", "鞋业", "有限公司", "有限责任公司", "公司"):
        if text.endswith(suffix):
            text = text[: -len(suffix)]
    return text


def _similarity(a: str, b: str) -> float:
    na, nb = _norm_name(a), _norm_name(b)
    if not na or not nb:
        return 0.0
    if na == nb:
        return 1.0
    if na in nb or nb in na:
        return 0.86

    def grams(s: str) -> set[str]:
        if len(s) < 2:
            return {s}
        return {s[i : i + 2] for i in range(len(s) - 1)}

    ga, gb = grams(na), grams(nb)
    inter = len(ga & gb)
    union = len(ga | gb) or 1
    return inter / union


def _match_customers(db: Session, tenant_id: int, hint: str | None) -> dict[str, Any]:
    hint = (hint or "").strip()
    partners = list(
        db.scalars(
            select(Partner).where(
                Partner.tenant_id == tenant_id,
                Partner.is_active.is_(True),
                (Partner.is_customer.is_(True)) | (Partner.is_brand.is_(True)),
            )
        ).all()
    )
    if not hint:
        return {
            "hint": "",
            "customer_id": None,
            "customer_name": "",
            "status": "needs_input",
            "candidates": [
                {"id": p.id, "name": p.short_name or p.name, "score": 0} for p in partners[:20]
            ],
        }

    scored: list[tuple[float, Partner]] = []
    for p in partners:
        names = [p.name or "", p.short_name or ""]
        score = max((_similarity(hint, n) for n in names if n), default=0.0)
        if score >= 0.55:
            scored.append((score, p))
    scored.sort(key=lambda x: (-x[0], x[1].id))
    candidates = [
        {"id": p.id, "name": p.short_name or p.name, "score": round(score, 3)}
        for score, p in scored[:8]
    ]

    unique_high = False
    if candidates and candidates[0]["score"] >= 0.92:
        if len(candidates) == 1 or candidates[0]["score"] - candidates[1]["score"] >= 0.08:
            unique_high = True
    if unique_high:
        c = candidates[0]
        return {
            "hint": hint,
            "customer_id": c["id"],
            "customer_name": c["name"],
            "status": "matched",
            "candidates": candidates,
        }

    return {
        "hint": hint,
        "customer_id": None,
        "customer_name": "",
        "status": "ambiguous",
        "candidates": candidates,
        "suggested_name": hint,
    }


def _match_product(db: Session, tenant_id: int, code: str) -> dict[str, Any]:
    code = (code or "").strip()
    products = list(
        db.scalars(
            select(OwnProduct).where(OwnProduct.tenant_id == tenant_id, OwnProduct.is_active.is_(True))
        ).all()
    )
    exact = next((p for p in products if p.product_code == code), None)
    if exact:
        return {
            "raw_code": code,
            "own_product_id": exact.id,
            "product_code": exact.product_code,
            "status": "matched",
            "candidates": [{"id": exact.id, "product_code": exact.product_code, "score": 1.0}],
        }

    scored: list[tuple[float, OwnProduct]] = []
    for p in products:
        score = _similarity(code, p.product_code or "")
        if score >= 0.6:
            scored.append((score, p))
    scored.sort(key=lambda x: (-x[0], x[1].id))
    candidates = [
        {"id": p.id, "product_code": p.product_code, "score": round(score, 3)}
        for score, p in scored[:6]
    ]
    if len(candidates) == 1 and candidates[0]["score"] >= 0.95:
        c = candidates[0]
        return {
            "raw_code": code,
            "own_product_id": c["id"],
            "product_code": c["product_code"],
            "status": "matched",
            "candidates": candidates,
        }
    return {
        "raw_code": code,
        "own_product_id": None,
        "product_code": None,
        "status": "missing" if not candidates else "ambiguous",
        "candidates": candidates,
    }


def _norm_attr(s: str) -> str:
    return re.sub(r"\s+", "", (s or "").strip().lower())


def _product_color_names(db: Session, tenant_id: int, product: OwnProduct) -> list[str]:
    color_ids = [c.color_id for c in (product.colors or [])]
    if not color_ids:
        return []
    colors = db.scalars(
        select(Color).where(Color.tenant_id == tenant_id, Color.id.in_(color_ids))
    ).all()
    return [c.name for c in colors if (c.name or "").strip()]


def _check_attr_vs_system(
    *,
    excel: str | None,
    system: str | None = None,
    system_list: list[str] | None = None,
    allow_empty_excel: bool = True,
) -> dict[str, Any]:
    excel_s = (excel or "").strip()
    systems = [s.strip() for s in (system_list or []) if (s or "").strip()]
    if system and not systems:
        sys_one = (system or "").strip()
        systems = [sys_one] if sys_one else []

    if not systems:
        return {
            "excel": excel_s or None,
            "system": None,
            "system_list": [],
            "match": True,
            "status": "no_system",
        }
    system_display = "、".join(systems)
    if not excel_s:
        if allow_empty_excel:
            return {
                "excel": None,
                "system": system_display,
                "system_list": systems,
                "match": True,
                "status": "no_excel",
            }
        return {
            "excel": None,
            "system": system_display,
            "system_list": systems,
            "match": False,
            "status": "mismatch",
        }
    ok = any(_norm_attr(excel_s) == _norm_attr(s) for s in systems)
    return {
        "excel": excel_s,
        "system": system_display,
        "system_list": systems,
        "match": ok,
        "status": "ok" if ok else "mismatch",
    }


def _enrich_line_attr_checks(
    db: Session,
    tenant_id: int,
    line: dict[str, Any],
    *,
    product_cache: dict[int, OwnProduct] | None = None,
) -> dict[str, Any]:
    """对照产品档案：颜色（可选色）、鞋面、内里/垫脚。"""
    pid = line.get("own_product_id")
    empty = {
        "color": {
            "excel": (line.get("color_name") or None),
            "system": None,
            "system_list": [],
            "match": True,
            "status": "no_product",
        },
        "fabric": {
            "excel": (line.get("fabric") or None),
            "system": None,
            "system_list": [],
            "match": True,
            "status": "no_product",
        },
        "lining": {
            "excel": (line.get("lining") or None),
            "system": None,
            "system_list": [],
            "match": True,
            "status": "no_product",
        },
    }
    if not pid:
        line["attr_checks"] = empty
        line["has_attr_mismatch"] = False
        return line

    cache = product_cache if product_cache is not None else {}
    product = cache.get(int(pid))
    if product is None:
        product = db.scalar(
            select(OwnProduct)
            .where(OwnProduct.id == int(pid), OwnProduct.tenant_id == tenant_id)
            .options(selectinload(OwnProduct.colors))
        )
        if product is not None:
            cache[int(pid)] = product
    if not product:
        line["attr_checks"] = empty
        line["has_attr_mismatch"] = False
        return line

    color_names = _product_color_names(db, tenant_id, product)
    checks = {
        "color": _check_attr_vs_system(
            excel=line.get("color_name"),
            system_list=color_names,
            allow_empty_excel=False,
        ),
        "fabric": _check_attr_vs_system(
            excel=line.get("fabric"),
            system=product.fabric,
            allow_empty_excel=True,
        ),
        "lining": _check_attr_vs_system(
            excel=line.get("lining"),
            system=product.lining,
            allow_empty_excel=True,
        ),
    }
    line["attr_checks"] = checks
    line["has_attr_mismatch"] = any(c.get("status") == "mismatch" for c in checks.values())
    line["system_fabric"] = (product.fabric or "").strip() or None
    line["system_lining"] = (product.lining or "").strip() or None
    line["system_colors"] = color_names
    return line


def _enrich_draft_attr_checks(db: Session, tenant_id: int, draft: dict[str, Any]) -> None:
    cache: dict[int, OwnProduct] = {}
    for ln in draft.get("lines") or []:
        if isinstance(ln, dict):
            _enrich_line_attr_checks(db, tenant_id, ln, product_cache=cache)


def _extract_images(content: bytes, session_id: str) -> list[dict[str, Any]]:
    images: list[dict[str, Any]] = []
    try:
        wb = load_workbook(BytesIO(content))
    except Exception:
        return images
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for img in getattr(ws, "_images", []) or []:
            try:
                data = img._data()  # noqa: SLF001
            except Exception:
                continue
            if not data:
                continue
            ext = ".png"
            try:
                fmt = (img.format or "").lower()
                if fmt in ("jpg", "jpeg"):
                    ext = ".jpg"
                elif fmt == "gif":
                    ext = ".gif"
                elif fmt == "webp":
                    ext = ".webp"
            except Exception:
                pass
            name = f"so_import_{session_id}_{len(images)}{ext}"
            dest = _uploads_dir() / name
            dest.write_bytes(data)
            images.append(
                {
                    "id": f"img_{len(images) + 1}",
                    "url": f"/uploads/{name}",
                    "sheet_name": sheet_name,
                    "role": IMAGE_ROLE_IGNORE,
                }
            )
            if len(images) >= 12:
                return images
    return images


def _ai_available() -> bool:
    return bool(get_settings().deepseek_api_key)


def _ai_map_headers(sample_rows: list[list[Any]]) -> dict[str, Any] | None:
    if not _ai_available():
        return None
    try:
        from langchain_openai import ChatOpenAI
    except Exception:
        return None

    s = get_settings()
    preview = []
    for row in sample_rows[:12]:
        preview.append([("" if c is None else str(c))[:40] for c in row[:20]])
    prompt = (
        "你是鞋厂订单 Excel 解析助手。根据下面若干行（含可能的表头），输出 JSON："
        '{"header_row_index":0,"customer_name":"","order_no":"","ordered_at":"YYYY-MM-DD或空",'
        '"delivery_date":"YYYY-MM-DD或空","notes":"做货要求文本或空",'
        '"confidence":0到1,"questions":["若不确定要问用户的问题"]}'
        "规则：不确定就降低 confidence 并写 questions，禁止编造数量。"
        f"\n表格预览:\n{json.dumps(preview, ensure_ascii=False)}"
    )
    try:
        llm = ChatOpenAI(
            api_key=s.deepseek_api_key,
            base_url=s.deepseek_base_url or "https://api.deepseek.com",
            model=s.deepseek_model or "deepseek-chat",
            temperature=0,
        )
        resp = llm.invoke(prompt)
        text = (resp.content or "").strip()
        m = re.search(r"\{[\s\S]*\}", text)
        if not m:
            return None
        data = json.loads(m.group(0))
        return data if isinstance(data, dict) else None
    except Exception:
        return None


def _sheet_rows_from_bytes(content: bytes) -> list[tuple[str, list[list[Any]]]]:
    wb = load_workbook(BytesIO(content), data_only=True)
    out: list[tuple[str, list[list[Any]]]] = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if any(any(c is not None and str(c).strip() for c in row) for row in rows):
            out.append((name, rows))
    return out


def _build_draft_from_parsed(db: Session, tenant_id: int, parsed: dict) -> dict[str, Any]:
    customer = _match_customers(db, tenant_id, parsed.get("customer_name"))
    lines = []
    for ln in parsed.get("lines") or []:
        prod = _match_product(db, tenant_id, ln.get("product_code") or "")
        delivery = parsed.get("delivery_date")
        lines.append(
            {
                "raw_product_code": ln.get("product_code"),
                "own_product_id": prod.get("own_product_id"),
                "product_code": prod.get("product_code"),
                "product_status": prod.get("status"),
                "product_candidates": prod.get("candidates") or [],
                "brand_name": ln.get("brand_name"),
                "customer_sku": ln.get("customer_sku"),
                "color_name": ln.get("color_name"),
                "fabric": ln.get("fabric"),
                "lining": ln.get("lining"),
                "unit_price": str(ln["unit_price"]) if ln.get("unit_price") is not None else None,
                "notes": ln.get("notes"),
                "items": ln.get("items") or [],
                "delivery_date": delivery.isoformat()
                if isinstance(delivery, date)
                else (delivery or None),
            }
        )
    ordered = parsed.get("ordered_at")
    delivery = parsed.get("delivery_date")
    draft = {
        "sheet_name": parsed.get("sheet_name"),
        "order_no": parsed.get("order_no") or "",
        "ordered_at": ordered.isoformat() if isinstance(ordered, date) else (ordered or ""),
        "delivery_date": delivery.isoformat() if isinstance(delivery, date) else (delivery or ""),
        "notes": parsed.get("notes") or "",
        "customer": customer,
        "lines": lines,
    }
    _enrich_draft_attr_checks(db, tenant_id, draft)
    return draft


def _collect_clarifications(draft: dict) -> list[dict[str, Any]]:
    qs: list[dict[str, Any]] = []
    cust = draft.get("customer") or {}
    if cust.get("status") in ("ambiguous", "needs_input"):
        qs.append(
            {
                "id": "customer",
                "type": "customer",
                "question": "请确认客户：从候选中选择，或手填客户名称",
                "candidates": cust.get("candidates") or [],
                "suggested_name": cust.get("suggested_name") or cust.get("hint") or "",
            }
        )
    for i, ln in enumerate(draft.get("lines") or []):
        st = ln.get("product_status")
        if st in ("ambiguous", "missing"):
            qs.append(
                {
                    "id": f"line_{i}_product",
                    "type": "product",
                    "line_index": i,
                    "question": (
                        f"第{i + 1}行工厂型号「{ln.get('raw_product_code') or ''}」"
                        "无法唯一匹配，请选择产品"
                    ),
                    "candidates": ln.get("product_candidates") or [],
                    "raw_code": ln.get("raw_product_code"),
                }
            )
    return qs


def _refresh_status(session: dict) -> None:
    clarifications = list(session.get("clarifications") or [])
    # 保留解析失败类澄清；其余按草稿重算
    parse_qs = [q for q in clarifications if q.get("type") in ("parse", "mapping")]
    draft_qs = _collect_clarifications(session.get("draft") or {})
    # 若已有草稿行，解析类问题可去掉
    if (session.get("draft") or {}).get("lines"):
        parse_qs = []
    session["clarifications"] = parse_qs + draft_qs
    if session.get("status") == STATUS_CONFIRMED:
        return
    if session.get("parse_error") and not (session.get("draft") or {}).get("lines"):
        if session["clarifications"]:
            session["status"] = STATUS_NEEDS_CLARIFICATION
        else:
            session["status"] = STATUS_FAILED
        return
    if session["clarifications"]:
        session["status"] = STATUS_NEEDS_CLARIFICATION
    else:
        session["status"] = STATUS_READY


def _public_session(session: dict) -> dict:
    return {
        "id": session["id"],
        "filename": session.get("filename"),
        "status": session.get("status"),
        "created_at": session.get("created_at"),
        "parse_error": session.get("parse_error"),
        "warnings": session.get("warnings") or [],
        "ai_used": bool(session.get("ai_used")),
        "draft": session.get("draft") or {},
        "images": session.get("images") or [],
        "clarifications": session.get("clarifications") or [],
        "result": session.get("result"),
        "can_confirm": session.get("status") == STATUS_READY
        and not (session.get("clarifications") or []),
    }


def create_import_session(
    db: Session,
    tenant_id: int,
    *,
    filename: str,
    content: bytes,
    created_by: int | None,
) -> dict:
    session_id = uuid.uuid4().hex
    tenant_dir = _tenant_dir(tenant_id)
    raw_path = tenant_dir / f"{session_id}.xlsx"
    raw_path.write_bytes(content)

    session: dict[str, Any] = {
        "id": session_id,
        "tenant_id": tenant_id,
        "created_by": created_by,
        "created_at": datetime.utcnow().isoformat(),
        "filename": filename,
        "file_path": str(raw_path),
        "status": STATUS_NEEDS_CLARIFICATION,
        "parse_error": None,
        "warnings": [],
        "ai_used": False,
        "draft": {},
        "images": _extract_images(content, session_id),
        "clarifications": [],
        "result": None,
    }

    try:
        parsed_sheets = parse_sales_order_workbook(content)
    except SalesOrderError as e:
        session["parse_error"] = e.message
        sheets = _sheet_rows_from_bytes(content)
        if sheets and _ai_available():
            _name, rows = sheets[0]
            mapped = _ai_map_headers(rows)
            session["ai_used"] = True
            if mapped:
                session["warnings"].append("规则解析失败，已用 AI 辅助查看表头，请改用标准模版或核对后重传")
                session["clarifications"] = [
                    {
                        "id": "ai_mapping_review",
                        "type": "mapping",
                        "question": (
                            "表头无法用规则稳定解析，不能自动生成明细数量。"
                            "请使用标准模版，或确认表头后重新上传。"
                            + (
                                (" 提示：" + "；".join(mapped.get("questions") or []))
                                if mapped.get("questions")
                                else ""
                            )
                        ),
                        "ai_mapping": mapped,
                    }
                ]
                session["status"] = STATUS_NEEDS_CLARIFICATION
                _save_session(tenant_id, session)
                return _public_session(session)
        session["clarifications"] = [
            {
                "id": "parse_failed",
                "type": "parse",
                "question": (
                    f"无法解析表格（{e.message}）。"
                    "请使用标准模版，或检查表头是否含「工厂型号」「颜色」。"
                ),
            }
        ]
        session["status"] = STATUS_FAILED
        _save_session(tenant_id, session)
        return _public_session(session)

    draft = _build_draft_from_parsed(db, tenant_id, parsed_sheets[0])
    if len(parsed_sheets) > 1:
        session["warnings"].append(
            f"文件含 {len(parsed_sheets)} 个工作表，本期仅导入第一张「{draft.get('sheet_name')}」"
        )
    session["draft"] = draft
    session["parse_error"] = None
    _refresh_status(session)
    _save_session(tenant_id, session)
    return _public_session(session)


def get_import_session(tenant_id: int, session_id: str) -> dict:
    return _public_session(_load_session(tenant_id, session_id))


def apply_clarifications(
    db: Session,
    tenant_id: int,
    session_id: str,
    answers: list[dict[str, Any]] | dict[str, Any],
) -> dict:
    session = _load_session(tenant_id, session_id)
    if session.get("status") == STATUS_CONFIRMED:
        raise SalesOrderError("import_already_confirmed", "该导入会话已确认建单")
    draft = session.get("draft") or {}
    if isinstance(answers, dict):
        answers = [
            {"id": k, **(v if isinstance(v, dict) else {"value": v})} for k, v in answers.items()
        ]

    for ans in answers:
        qid = str(ans.get("id") or "")
        if qid == "customer" or ans.get("type") == "customer":
            cust = draft.get("customer") or {}
            cid = ans.get("customer_id")
            cname = (ans.get("customer_name") or ans.get("value") or "").strip()
            if cid:
                p = db.get(Partner, int(cid))
                if not p or p.tenant_id != tenant_id or not p.is_active:
                    raise SalesOrderError("invalid_customer", "客户不存在或未启用")
                cust["customer_id"] = p.id
                cust["customer_name"] = cname or p.short_name or p.name
                cust["status"] = "matched"
            elif cname:
                cust["customer_id"] = None
                cust["customer_name"] = cname
                cust["status"] = "matched"
            else:
                raise SalesOrderError("customer_required", "请选择或填写客户")
            draft["customer"] = cust
            continue

        if qid.startswith("line_") and qid.endswith("_product"):
            try:
                idx = int(qid.split("_")[1])
            except Exception as e:
                raise SalesOrderError("bad_clarify", f"无效澄清项 {qid}") from e
            lines = draft.get("lines") or []
            if idx < 0 or idx >= len(lines):
                raise SalesOrderError("bad_clarify", f"明细行不存在：{idx}")
            pid = ans.get("own_product_id") or ans.get("value")
            if not pid:
                raise SalesOrderError("product_required", "请选择产品")
            p = db.get(OwnProduct, int(pid))
            if not p or p.tenant_id != tenant_id or not p.is_active:
                raise SalesOrderError("invalid_product", "产品不存在或未启用")
            lines[idx]["own_product_id"] = p.id
            lines[idx]["product_code"] = p.product_code
            lines[idx]["product_status"] = "matched"
            draft["lines"] = lines

    session["draft"] = draft
    # 用户已处理解析类问题时，清掉 parse/mapping 澄清
    session["clarifications"] = [
        q
        for q in (session.get("clarifications") or [])
        if q.get("type") not in ("parse", "mapping")
    ]
    _enrich_draft_attr_checks(db, tenant_id, draft)
    _refresh_status(session)
    _save_session(tenant_id, session)
    return _public_session(session)


def patch_import_draft(
    db: Session,
    tenant_id: int,
    session_id: str,
    patch: dict[str, Any],
) -> dict:
    session = _load_session(tenant_id, session_id)
    if session.get("status") == STATUS_CONFIRMED:
        raise SalesOrderError("import_already_confirmed", "该导入会话已确认建单")
    draft = session.get("draft") or {}

    for key in ("order_no", "ordered_at", "delivery_date", "notes"):
        if key in patch:
            draft[key] = patch.get(key) or ""

    if "customer" in patch and isinstance(patch["customer"], dict):
        c = patch["customer"]
        cust = draft.get("customer") or {}
        if "customer_id" in c:
            pid = int(c["customer_id"]) if c.get("customer_id") else None
            if pid:
                p = db.get(Partner, pid)
                if not p or p.tenant_id != tenant_id:
                    raise SalesOrderError("invalid_customer", "客户不存在")
                cust["customer_id"] = p.id
                cust["customer_name"] = (
                    c.get("customer_name") or p.short_name or p.name or ""
                ).strip()
                cust["status"] = "matched"
            else:
                cust["customer_id"] = None
        if c.get("customer_name") is not None:
            name = (c.get("customer_name") or "").strip()
            cust["customer_name"] = name
            if name:
                cust["status"] = "matched"
            elif not cust.get("customer_id"):
                cust["status"] = "needs_input"
        draft["customer"] = cust

    if "lines" in patch and isinstance(patch["lines"], list):
        lines = draft.get("lines") or []
        for item in patch["lines"]:
            if not isinstance(item, dict) or "index" not in item:
                continue
            idx = int(item["index"])
            if idx < 0 or idx >= len(lines):
                continue
            row = lines[idx]
            if item.get("own_product_id"):
                p = db.get(OwnProduct, int(item["own_product_id"]))
                if not p or p.tenant_id != tenant_id:
                    raise SalesOrderError("invalid_product", "产品不存在")
                row["own_product_id"] = p.id
                row["product_code"] = p.product_code
                row["product_status"] = "matched"
            for k in ("color_name", "fabric", "lining", "brand_name", "customer_sku", "notes", "unit_price"):
                if k in item:
                    row[k] = item.get(k)
            lines[idx] = row
        draft["lines"] = lines

    if "images" in patch and isinstance(patch["images"], list):
        role_by_id = {
            str(it.get("id")): it.get("role")
            for it in patch["images"]
            if isinstance(it, dict) and it.get("id")
        }
        for img in session.get("images") or []:
            role = role_by_id.get(img["id"])
            if role in (IMAGE_ROLE_IGNORE, IMAGE_ROLE_LOGO, IMAGE_ROLE_NOTES):
                img["role"] = role
        seen: dict[str, str] = {}
        for img in session.get("images") or []:
            role = img.get("role") or IMAGE_ROLE_IGNORE
            if role in (IMAGE_ROLE_LOGO, IMAGE_ROLE_NOTES):
                if role in seen and seen[role] != img["id"]:
                    img["role"] = IMAGE_ROLE_IGNORE
                else:
                    seen[role] = img["id"]

    session["draft"] = draft
    _enrich_draft_attr_checks(db, tenant_id, draft)
    _refresh_status(session)
    _save_session(tenant_id, session)
    return _public_session(session)


def confirm_import_session(
    db: Session,
    tenant_id: int,
    session_id: str,
    *,
    created_by: int | None,
) -> dict:
    session = _load_session(tenant_id, session_id)
    if session.get("status") == STATUS_CONFIRMED and session.get("result"):
        return _public_session(session)

    _refresh_status(session)
    if session.get("clarifications"):
        raise SalesOrderError("needs_clarification", "仍有待澄清项，请先确认客户/产品后再导入")
    draft = session.get("draft") or {}
    cust = draft.get("customer") or {}
    customer_name = (cust.get("customer_name") or "").strip()
    customer_id = cust.get("customer_id")
    if not customer_name and not customer_id:
        raise SalesOrderError("customer_required", "请确认客户")

    order_no = (draft.get("order_no") or "").strip()
    if order_no and db.scalar(
        select(SalesOrder).where(SalesOrder.tenant_id == tenant_id, SalesOrder.order_no == order_no)
    ):
        raise SalesOrderError("duplicate_order_no", f"销售订单号已存在: {order_no}")

    color_cache: dict[str, Color] = {
        c.name: c for c in db.scalars(select(Color).where(Color.tenant_id == tenant_id)).all()
    }
    size_cache: dict[str, Size] = {
        s.size_value: s for s in db.scalars(select(Size).where(Size.tenant_id == tenant_id)).all()
    }

    delivery_raw = (draft.get("delivery_date") or "").strip()
    delivery: date | None = None
    if delivery_raw:
        try:
            delivery = date.fromisoformat(delivery_raw[:10])
        except ValueError as e:
            raise SalesOrderError("bad_delivery_date", f"出货日期无效: {delivery_raw}") from e

    line_ins: list[SalesOrderLineIn] = []
    for i, ln in enumerate(draft.get("lines") or []):
        pid = ln.get("own_product_id")
        if not pid:
            raise SalesOrderError("product_required", f"第{i + 1}行未选择产品")
        product = db.get(OwnProduct, int(pid))
        if not product or product.tenant_id != tenant_id:
            raise SalesOrderError("invalid_product", f"第{i + 1}行产品无效")
        color_name = (ln.get("color_name") or "").strip()
        if not color_name:
            raise SalesOrderError("missing_color", f"第{i + 1}行缺少颜色")
        color = _get_or_create_color(db, tenant_id, color_name, color_cache)
        _ensure_product_color(db, tenant_id, product, color.id)
        items: list[SalesOrderLineItemIn] = []
        for it in ln.get("items") or []:
            size = _get_or_create_size(db, tenant_id, str(it.get("size_value") or ""), size_cache)
            items.append(SalesOrderLineItemIn(size_id=size.id, qty=int(it["qty"])))
        if not items:
            raise SalesOrderError("empty_items", f"第{i + 1}行无有效码数")
        unit_price = None
        if ln.get("unit_price") not in (None, ""):
            unit_price = Decimal(str(ln["unit_price"]))
        line_ins.append(
            SalesOrderLineIn(
                own_product_id=product.id,
                color_id=color.id,
                fabric=(ln.get("fabric") or None),
                lining=(ln.get("lining") or None),
                customer_sku=(ln.get("customer_sku") or None),
                brand_name=(ln.get("brand_name") or None),
                delivery_date=delivery,
                unit_price=unit_price,
                notes=(ln.get("notes") or None),
                items=items,
            )
        )
    if not line_ins:
        raise SalesOrderError("empty_lines", "没有可导入的明细")

    brand_logo_url = None
    notes_image_url = None
    for img in session.get("images") or []:
        if img.get("role") == IMAGE_ROLE_LOGO:
            brand_logo_url = img.get("url")
        elif img.get("role") == IMAGE_ROLE_NOTES:
            notes_image_url = img.get("url")

    ordered_raw = (draft.get("ordered_at") or "").strip()
    ordered_at = date.today()
    if ordered_raw:
        try:
            ordered_at = date.fromisoformat(ordered_raw[:10])
        except ValueError as e:
            raise SalesOrderError("bad_ordered_at", f"下单日期无效: {ordered_raw}") from e

    so = create_sales_order(
        db,
        tenant_id,
        SalesOrderCreate(
            order_no=order_no or None,
            customer_id=int(customer_id) if customer_id else None,
            customer_name=customer_name or None,
            ordered_at=ordered_at,
            notes=(draft.get("notes") or None),
            brand_logo_url=brand_logo_url,
            notes_image_url=notes_image_url,
            lines=line_ins,
        ),
        created_by=created_by,
    )
    session["status"] = STATUS_CONFIRMED
    session["clarifications"] = []
    session["result"] = {
        "sales_order_id": so.id,
        "order_no": so.order_no,
        "order": serialize_sales_order(db, tenant_id, so),
    }
    _save_session(tenant_id, session)
    return _public_session(session)
