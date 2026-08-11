"""采购单 Excel 导出：版式对齐打印预览（供应商联 / 可选内部联）。"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

import qrcode
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

INK = "111111"
MUTED = "555555"
HEADER_BG = "F3F4F6"
LINE = "333333"
DRAFT = "C45656"

thin_side = Side(style="thin", color=LINE)
thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
money_fmt = "0.00"


def _font(bold: bool = False, size: int = 11, color: str = INK) -> Font:
    return Font(name="微软雅黑", bold=bold, size=size, color=color)


def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _set_widths(ws: Worksheet, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _txt(v: Any, default: str = "—") -> str:
    if v is None:
        return default
    s = str(v).strip()
    return s if s else default


def _date_txt(v: Any) -> str:
    if v is None:
        return "—"
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).replace("T", " ")
    return s[:10] if s else "—"


def _num(v: Any) -> float | None:
    if v is None:
        return None
    try:
        return float(Decimal(str(v)))
    except Exception:
        return None


def _qr_image(url: str, size_px: int = 120) -> XLImage | None:
    if not url:
        return None
    try:
        img = qrcode.make(url)
        buf = BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        xl = XLImage(buf)
        xl.width = size_px
        xl.height = size_px
        return xl
    except Exception:
        return None


def _write_supplier_sheet(
    ws: Worksheet,
    detail: dict,
    *,
    public_url: str | None,
) -> None:
    ws.sheet_view.showGridLines = False
    _set_widths(ws, {1: 16, 2: 28, 3: 10, 4: 12, 5: 12, 6: 14})

    is_draft = str(detail.get("status") or "") == "draft"
    row = 1
    if is_draft:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row, 1, "【待下单】未正式下单，仅供内部参考")
        cell.font = _font(True, 12, DRAFT)
        cell.alignment = _align("center")
        row = 2

    # 买方抬头 + 二维码
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row, 1, _txt(detail.get("buyer_name")))
    c.font = _font(True, 16)
    c.alignment = _align("left", "center")
    ws.row_dimensions[row].height = 24
    buyer_row = row

    qr = _qr_image(public_url or "")
    if qr:
        ws.add_image(qr, f"F{row}")
        ws.row_dimensions[row].height = 20

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    contact = (
        f"联系人{_txt(detail.get('buyer_contact_person'))}    "
        f"电话{_txt(detail.get('buyer_contact_mobile'))}"
    )
    c = ws.cell(row, 1, contact)
    c.font = _font(False, 10)
    c.alignment = _align("left")

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row, 1, f"地址{_txt(detail.get('buyer_address'))}")
    c.font = _font(False, 10)
    c.alignment = _align("left", wrap=True)

    # 单号标在二维码下方
    po_label_row = buyer_row + 3
    if qr:
        c = ws.cell(po_label_row, 6, _txt(detail.get("po_no"), ""))
        c.font = _font(False, 9, MUTED)
        c.alignment = _align("center")

    row = max(row + 2, po_label_row + 2)

    meta = [
        (f"采购单号：{_txt(detail.get('po_no'))}", f"协商交货日期：{_txt(detail.get('expected_date'))}"),
        (
            f"供应商：{_txt(detail.get('partner_name'))}",
            f"下单日期：{_date_txt(detail.get('ordered_at'))}",
        ),
        (
            f"联系人：{_txt(detail.get('partner_contact_name'))}  "
            f"{_txt(detail.get('partner_contact_mobile'), '')}".rstrip(),
            "",
        ),
    ]
    for left, right in meta:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row, 1, left)
        c.font = _font(False, 11)
        if right:
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
            c2 = ws.cell(row, 4, right)
            c2.font = _font(False, 11)
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row, 1, f"供应商地址：{_txt(detail.get('partner_address'))}")
    c.font = _font(False, 11)
    c.alignment = _align("left", wrap=True)
    row += 2

    headers = ["物料编码", "名称", "单位", "数量", "单价", "金额"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row, col, h)
        cell.font = _font(True, 11)
        cell.fill = _fill(HEADER_BG)
        cell.border = thin
        cell.alignment = _align("right" if col >= 4 else "left")
    header_row = row
    row += 1

    for ln in detail.get("summary_lines") or []:
        vals = [
            _txt(ln.get("supplier_product_code"), ""),
            _txt(ln.get("supplier_product_name")),
            _txt(ln.get("pricing_unit_name")),
            _num(ln.get("qty")),
            _num(ln.get("unit_price")),
            _num(ln.get("amount")),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row, col, val if val is not None else "—")
            cell.font = _font(False, 11)
            cell.border = thin
            if col >= 4:
                cell.alignment = _align("right")
                if isinstance(val, (int, float)) and col >= 5:
                    cell.number_format = money_fmt
                elif isinstance(val, (int, float)):
                    cell.number_format = "0.####"
            else:
                cell.alignment = _align("left")
        row += 1

    if not (detail.get("summary_lines") or []):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row, 1, "（无明细）")
        cell.border = thin
        cell.alignment = _align("center")
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    total = _num(detail.get("summary_total_amount")) or 0
    cell = ws.cell(row, 1, f"合计金额：¥{total:.2f}")
    cell.font = _font(True, 12)
    cell.alignment = _align("right")
    row += 2

    note = f"备注：{_txt(detail.get('notes'), '无')}"
    tax = _txt(detail.get("tax_note"), "")
    if tax and tax != "—":
        note = f"{note}\n{tax}"
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=6)
    cell = ws.cell(row, 1, note)
    cell.font = _font(False, 10, MUTED)
    cell.alignment = _align("left", "top", wrap=True)
    ws.row_dimensions[row].height = 18
    ws.row_dimensions[row + 1].height = 18
    row += 3

    # 签字区
    ws.cell(row, 1, "采购方签字/盖章").font = _font(False, 10, MUTED)
    ws.cell(row, 4, "供应商签字/盖章").font = _font(False, 10, MUTED)
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row, 1, "________________").font = _font(False, 11)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
    ws.cell(row, 4, "________________").font = _font(False, 11)
    row += 1
    ws.cell(row, 1, "日期：________").font = _font(False, 10, MUTED)
    ws.cell(row, 4, "日期：________").font = _font(False, 10, MUTED)

    # 打印区域大致覆盖内容
    ws.print_title_rows = f"1:{header_row}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def _write_internal_sheet(ws: Worksheet, detail: dict) -> None:
    ws.sheet_view.showGridLines = False
    _set_widths(ws, {1: 16, 2: 28, 3: 10, 4: 14, 5: 12, 6: 12})

    ws.merge_cells("A1:F1")
    c = ws.cell(1, 1, f"采购单内部联 · {_txt(detail.get('po_no'))}")
    c.font = _font(True, 14)
    c.alignment = _align("left")

    ws.merge_cells("A2:F2")
    c = ws.cell(2, 1, "以下为分订单明细，仅内部跟单/到货回写使用，勿发给供应商。")
    c.font = _font(False, 10, MUTED)

    headers = ["物料编码", "名称", "单位", "订单", "数量", "单价"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(4, col, h)
        cell.font = _font(True, 11)
        cell.fill = _fill(HEADER_BG)
        cell.border = thin
        cell.alignment = _align("right" if col >= 5 else "left")

    row = 5
    for ln in detail.get("lines") or []:
        vals = [
            _txt(ln.get("supplier_product_code"), ""),
            _txt(ln.get("supplier_product_name")),
            _txt(ln.get("pricing_unit_name")),
            _txt(ln.get("order_no")),
            _num(ln.get("qty")),
            _num(ln.get("unit_price")),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row, col, val if val is not None else "—")
            cell.font = _font(False, 11)
            cell.border = thin
            if col >= 5:
                cell.alignment = _align("right")
                if isinstance(val, (int, float)) and col == 6:
                    cell.number_format = money_fmt
                elif isinstance(val, (int, float)):
                    cell.number_format = "0.####"
            else:
                cell.alignment = _align("left")
        row += 1

    if not (detail.get("lines") or []):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row, 1, "（无明细）")
        cell.border = thin
        cell.alignment = _align("center")


def build_purchase_order_workbook(
    detail: dict,
    *,
    include_internal: bool = False,
    public_url: str | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "供应商联"
    _write_supplier_sheet(ws, detail, public_url=public_url)

    if include_internal:
        ws2 = wb.create_sheet("内部联")
        _write_internal_sheet(ws2, detail)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
