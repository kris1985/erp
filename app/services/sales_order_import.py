"""销售订单 Excel 导入（订单头 + 明细同行）。

兼容客户下单模版：表头含「工厂型号 / 品牌 / 客人型号 / 颜色 / 面料 / 内里/垫脚 / 码数列…」。
同文件多 sheet 各建一单；同工厂型号跨行合并色别时，空型号向前填充。
"""

from __future__ import annotations

import re
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.models import Color, OwnProduct, OwnProductColor, SalesOrder, Size
from app.schemas.api import SalesOrderCreate, SalesOrderLineIn, SalesOrderLineItemIn
from app.services.sales_order_service import SalesOrderError, create_sales_order

HEADER_FACTORY = "工厂型号"
HEADER_BRAND = "品牌"
HEADER_CUSTOMER_SKU_ALIASES = ("客人型号", "客户型号")
HEADER_COLOR = "颜色"
HEADER_FABRIC = ("面料", "鞋面")
HEADER_LINING = ("内里/垫脚", "内里", "垫脚")
HEADER_QTY_MARKERS = ("件数", "总数", "单价", "金额", "备注")
STOP_MARKERS = ("合计", "做货要求")


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _parse_cn_date(text: str) -> date | None:
    raw = (text or "").strip()
    if not raw:
        return None
    m = re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", raw)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(raw[:19], fmt).date()
        except ValueError:
            continue
    if isinstance(text, datetime):
        return text.date()
    return None


def _parse_labeled_date(cell: Any) -> date | None:
    if isinstance(cell, datetime):
        return cell.date()
    if isinstance(cell, date):
        return cell
    return _parse_cn_date(_cell_str(cell))


def _to_int_qty(v: Any) -> int | None:
    if v is None or v == "":
        return None
    try:
        n = int(float(v))
    except (TypeError, ValueError):
        return None
    return n if n > 0 else None


def _to_decimal(v: Any) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        d = Decimal(str(v).strip())
    except (InvalidOperation, ValueError):
        return None
    return d


def _find_header_row(rows: list[list[Any]]) -> tuple[int, dict[str, int], list[tuple[int, str]]]:
    """返回 (header_row_idx, named_cols, size_cols[(col_idx, size_value)])."""
    for idx, row in enumerate(rows):
        labels = {_cell_str(c): i for i, c in enumerate(row) if _cell_str(c)}
        if HEADER_FACTORY not in labels or HEADER_COLOR not in labels:
            continue
        named: dict[str, int] = {
            "factory": labels[HEADER_FACTORY],
            "color": labels[HEADER_COLOR],
        }
        if HEADER_BRAND in labels:
            named["brand"] = labels[HEADER_BRAND]
        for name in HEADER_CUSTOMER_SKU_ALIASES:
            if name in labels:
                named["customer_sku"] = labels[name]
                break
        for name in HEADER_FABRIC:
            if name in labels:
                named["fabric"] = labels[name]
                break
        for name in HEADER_LINING:
            if name in labels:
                named["lining"] = labels[name]
                break
        for name in ("单价",):
            if name in labels:
                named["unit_price"] = labels[name]
        for name in ("备注",):
            if name in labels:
                named["notes"] = labels[name]

        size_start = named["color"] + 1
        if "lining" in named:
            size_start = max(size_start, named["lining"] + 1)
        elif "fabric" in named:
            size_start = max(size_start, named["fabric"] + 1)

        stop_cols = [
            labels[m] for m in HEADER_QTY_MARKERS if m in labels
        ]
        size_end = min(stop_cols) if stop_cols else len(row)

        size_cols: list[tuple[int, str]] = []
        for col_i in range(size_start, size_end):
            label = _cell_str(row[col_i])
            if not label:
                continue
            # 尺码通常是数字，也兼容 35.5 等
            if re.fullmatch(r"\d+(\.\d+)?", label) or re.fullmatch(r"[A-Za-z]+", label):
                size_cols.append((col_i, label))
        if not size_cols:
            raise SalesOrderError("bad_xlsx", f"第{idx + 1}行表头未识别到码数列")
        return idx, named, size_cols
    raise SalesOrderError("bad_xlsx", "未找到表头（需含「工厂型号」「颜色」）")


def _extract_meta(rows: list[list[Any]], header_idx: int) -> dict[str, Any]:
    order_no = ""
    customer_name = ""
    ordered_at: date | None = None
    delivery_date: date | None = None
    for row in rows[:header_idx]:
        for cell in row:
            text = _cell_str(cell)
            if not text:
                continue
            if not order_no:
                m = re.search(r"订单号\s*[:：]\s*(\S+)", text)
                if m:
                    order_no = m.group(1).strip()
            if ordered_at is None and "下单时间" in text:
                ordered_at = _parse_labeled_date(text.split("下单时间", 1)[-1].lstrip(" :："))
                if ordered_at is None:
                    ordered_at = _parse_labeled_date(text)
            if delivery_date is None and ("出货时间" in text or "交货" in text):
                delivery_date = _parse_labeled_date(
                    re.split(r"出货时间|交货日期|交货时间", text, maxsplit=1)[-1].lstrip(" :：")
                )
                if delivery_date is None:
                    delivery_date = _parse_labeled_date(text)
            if not customer_name and text.endswith("订单明细"):
                customer_name = text[: -len("订单明细")].strip()
            if not customer_name:
                m = re.match(r"客户\s*[:：]\s*(.+)$", text)
                if m:
                    customer_name = m.group(1).strip()
    return {
        "order_no": order_no,
        "customer_name": customer_name,
        "ordered_at": ordered_at,
        "delivery_date": delivery_date,
    }


def _extract_notes(rows: list[list[Any]], start_idx: int) -> str | None:
    notes: list[str] = []
    started = False
    for row in rows[start_idx:]:
        joined = " ".join(_cell_str(c) for c in row if _cell_str(c)).strip()
        if not joined:
            if started:
                break
            continue
        if "做货要求" in joined:
            started = True
            rest = re.split(r"做货要求\s*[:：]?", joined, maxsplit=1)[-1].strip()
            if rest:
                notes.append(rest)
            continue
        if started:
            notes.append(joined)
    text = "\n".join(notes).strip()
    return text or None


def parse_sales_order_sheet(rows: list[list[Any]]) -> dict[str, Any]:
    header_idx, named, size_cols = _find_header_row(rows)
    meta = _extract_meta(rows, header_idx)
    lines: list[dict[str, Any]] = []
    last_factory = ""
    last_brand = ""
    last_sku = ""
    data_end = header_idx + 1
    for ridx in range(header_idx + 1, len(rows)):
        row = rows[ridx]
        first = _cell_str(row[0]) if row else ""
        if any(first.startswith(m) for m in STOP_MARKERS) or first in STOP_MARKERS:
            data_end = ridx
            break
        # 空行跳过
        if not any(_cell_str(c) for c in row):
            continue
        factory = _cell_str(row[named["factory"]]) if named["factory"] < len(row) else ""
        if factory:
            last_factory = factory
        else:
            factory = last_factory
        if not factory:
            continue
        color = _cell_str(row[named["color"]]) if named["color"] < len(row) else ""
        if not color:
            continue

        brand = ""
        if "brand" in named and named["brand"] < len(row):
            brand = _cell_str(row[named["brand"]])
        if brand:
            last_brand = brand
        else:
            brand = last_brand

        sku = ""
        if "customer_sku" in named and named["customer_sku"] < len(row):
            sku = _cell_str(row[named["customer_sku"]])
        if sku:
            last_sku = sku
        else:
            sku = last_sku

        fabric = ""
        if "fabric" in named and named["fabric"] < len(row):
            fabric = _cell_str(row[named["fabric"]])
        lining = ""
        if "lining" in named and named["lining"] < len(row):
            lining = _cell_str(row[named["lining"]])
        unit_price = None
        if "unit_price" in named and named["unit_price"] < len(row):
            unit_price = _to_decimal(row[named["unit_price"]])
        notes = ""
        if "notes" in named and named["notes"] < len(row):
            notes = _cell_str(row[named["notes"]])

        items: list[dict[str, Any]] = []
        for col_i, size_value in size_cols:
            if col_i >= len(row):
                continue
            qty = _to_int_qty(row[col_i])
            if qty:
                items.append({"size_value": size_value, "qty": qty})
        if not items:
            continue
        lines.append(
            {
                "product_code": factory,
                "brand_name": brand or None,
                "customer_sku": sku or None,
                "color_name": color,
                "fabric": fabric or None,
                "lining": lining or None,
                "unit_price": unit_price,
                "notes": notes or None,
                "items": items,
            }
        )
        data_end = ridx + 1

    if not lines:
        raise SalesOrderError("bad_xlsx", "未解析到订单明细行")

    notes = _extract_notes(rows, data_end)
    # 若表尾未标「做货要求」，再扫全表兜底
    if not notes:
        notes = _extract_notes(rows, header_idx + 1)

    return {
        **meta,
        "notes": notes,
        "lines": lines,
    }


def _sheet_to_rows(ws) -> list[list[Any]]:
    return [list(row) for row in ws.iter_rows(values_only=True)]


def parse_sales_order_workbook(content: bytes) -> list[dict[str, Any]]:
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        raise SalesOrderError("bad_xlsx", f"无法读取 Excel：{e}") from e
    parsed: list[dict[str, Any]] = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = _sheet_to_rows(ws)
        if not any(any(_cell_str(c) for c in row) for row in rows):
            continue
        try:
            data = parse_sales_order_sheet(rows)
        except SalesOrderError:
            # 跳过无表头的说明 sheet
            continue
        data["sheet_name"] = name
        parsed.append(data)
    if not parsed:
        raise SalesOrderError("bad_xlsx", "未从 Excel 中解析到可导入的订单")
    return parsed


def _get_or_create_color(db: Session, tenant_id: int, name: str, cache: dict[str, Color]) -> Color:
    key = name.strip()
    if key in cache:
        return cache[key]
    color = db.scalar(select(Color).where(Color.tenant_id == tenant_id, Color.name == key))
    if not color:
        code = f"C{uuid.uuid4().hex[:6].upper()}"
        while db.scalar(select(Color).where(Color.tenant_id == tenant_id, Color.code == code)):
            code = f"C{uuid.uuid4().hex[:6].upper()}"
        color = Color(tenant_id=tenant_id, name=key, code=code)
        db.add(color)
        db.flush()
    cache[key] = color
    return color


def _get_or_create_size(db: Session, tenant_id: int, size_value: str, cache: dict[str, Size]) -> Size:
    key = size_value.strip()
    if key in cache:
        return cache[key]
    size = db.scalar(select(Size).where(Size.tenant_id == tenant_id, Size.size_value == key))
    if not size:
        size = Size(tenant_id=tenant_id, size_value=key, sort_order=0, is_active=True)
        db.add(size)
        db.flush()
    elif getattr(size, "is_active", True) is False:
        size.is_active = True
        db.flush()
    cache[key] = size
    return size


def _ensure_product_color(
    db: Session, tenant_id: int, product: OwnProduct, color_id: int
) -> None:
    product = db.scalar(
        select(OwnProduct)
        .where(OwnProduct.id == product.id, OwnProduct.tenant_id == tenant_id)
        .options(selectinload(OwnProduct.colors))
    )
    if not product:
        raise SalesOrderError("invalid_product", "产品不存在")
    allowed = {c.color_id for c in (product.colors or [])}
    if color_id in allowed:
        return
    if allowed:
        raise SalesOrderError(
            "invalid_color",
            f"颜色与产品「{product.product_code}」不匹配。一色一款请用对应货号，不要往已绑色的产品上补色",
        )
    # 旧货号未绑色：按本单颜色补一条，便于后续下单校验
    db.add(
        OwnProductColor(
            tenant_id=tenant_id,
            own_product_id=product.id,
            color_id=color_id,
        )
    )
    db.flush()


def import_sales_orders_xlsx(
    db: Session,
    tenant_id: int,
    content: bytes,
    *,
    created_by: int | None,
    customer_name_override: str | None = None,
) -> dict:
    sheets = parse_sales_order_workbook(content)
    products = {
        p.product_code: p
        for p in db.scalars(
            select(OwnProduct).where(OwnProduct.tenant_id == tenant_id, OwnProduct.is_active.is_(True))
        ).all()
    }
    color_cache: dict[str, Color] = {
        c.name: c for c in db.scalars(select(Color).where(Color.tenant_id == tenant_id)).all()
    }
    size_cache: dict[str, Size] = {
        s.size_value: s for s in db.scalars(select(Size).where(Size.tenant_id == tenant_id)).all()
    }

    created: list[str] = []
    skipped: list[str] = []
    errors: list[str] = []

    for sheet in sheets:
        sheet_label = sheet.get("sheet_name") or "Sheet"
        order_no = (sheet.get("order_no") or "").strip()
        customer_name = (
            (customer_name_override or "").strip()
            or (sheet.get("customer_name") or "").strip()
        )
        if not customer_name:
            errors.append(f"「{sheet_label}」：无法识别客户名（标题需如「某某鞋业订单明细」或填写客户）")
            continue
        if order_no and db.scalar(
            select(SalesOrder).where(SalesOrder.tenant_id == tenant_id, SalesOrder.order_no == order_no)
        ):
            skipped.append(order_no)
            continue

        line_ins: list[SalesOrderLineIn] = []
        sheet_errors: list[str] = []
        delivery = sheet.get("delivery_date")
        for i, ln in enumerate(sheet["lines"], start=1):
            product = products.get(ln["product_code"])
            if not product:
                sheet_errors.append(f"明细{i}：找不到工厂型号 {ln['product_code']}")
                continue
            color = _get_or_create_color(db, tenant_id, ln["color_name"], color_cache)
            try:
                _ensure_product_color(db, tenant_id, product, color.id)
            except SalesOrderError as e:
                sheet_errors.append(f"明细{i}：{e.message}")
                continue
            items: list[SalesOrderLineItemIn] = []
            for it in ln["items"]:
                size = _get_or_create_size(db, tenant_id, it["size_value"], size_cache)
                items.append(SalesOrderLineItemIn(size_id=size.id, qty=int(it["qty"])))
            if not items:
                sheet_errors.append(f"明细{i}：无有效码数数量")
                continue
            line_ins.append(
                SalesOrderLineIn(
                    own_product_id=product.id,
                    color_id=color.id,
                    fabric=ln.get("fabric"),
                    lining=ln.get("lining"),
                    customer_sku=ln.get("customer_sku"),
                    brand_name=ln.get("brand_name"),
                    delivery_date=delivery,
                    unit_price=ln.get("unit_price"),
                    notes=ln.get("notes"),
                    items=items,
                )
            )

        if sheet_errors:
            errors.extend([f"「{sheet_label}」{e}" for e in sheet_errors])
            continue
        if not line_ins:
            errors.append(f"「{sheet_label}」：没有可导入的明细")
            continue

        try:
            so = create_sales_order(
                db,
                tenant_id,
                SalesOrderCreate(
                    order_no=order_no or None,
                    customer_name=customer_name,
                    ordered_at=sheet.get("ordered_at") or date.today(),
                    notes=sheet.get("notes"),
                    lines=line_ins,
                ),
                created_by=created_by,
            )
            created.append(so.order_no)
        except SalesOrderError as e:
            errors.append(f"「{sheet_label}」：{e.message}")
        except Exception as e:
            errors.append(f"「{sheet_label}」：导入失败 {e}")

    return {
        "created": created,
        "created_count": len(created),
        "skipped": skipped,
        "skipped_count": len(skipped),
        "errors": errors,
        "message": (
            f"导入完成：新建 {len(created)} 单"
            + (f"，跳过已存在 {len(skipped)} 单" if skipped else "")
            + (f"，错误 {len(errors)} 条" if errors else "")
        ),
    }


def build_sales_order_import_template_bytes() -> bytes:
    """生成精简模版（无图片），结构与客户订单明细一致。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "订单明细"
    ws["A1"] = "示例客户订单明细"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:R1")
    ws["A3"] = "TO:  工厂名称"
    ws["L3"] = "下单时间：2026年 8 月 10 日"
    ws["A4"] = "订单号：SO示例001"
    ws["L4"] = "出货时间：2026年 8 月 25 日"

    headers = [
        "工厂型号",
        "品牌",
        "客人型号",
        "图片",
        "颜色",
        "面料",
        "内里/垫脚",
        35,
        36,
        37,
        38,
        39,
        40,
        "件数",
        "总数",
        "单价",
        "金额",
        "备注",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(5, col, h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    ws.append(
        [
            "OP-001",
            "示例品牌",
            "CUS-001",
            None,
            "黑色",
            "牛皮",
            "猪皮",
            None,
            2,
            5,
            4,
            3,
            1,
            None,
            15,
            88,
            None,
            "",
        ]
    )
    ws.append(
        [
            None,
            None,
            None,
            None,
            "米色",
            "牛皮",
            "猪皮",
            None,
            2,
            3,
            3,
            3,
            1,
            None,
            12,
            88,
            None,
            "",
        ]
    )
    ws.append(["合计："])
    ws.append(["做货要求："])
    ws.append([None, "1.LOGO 按客人要求"])
    ws.append([None, "2.品牌包装"])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def packaged_template_path() -> Path | None:
    path = Path(__file__).resolve().parents[2] / "assets" / "templates" / "sales_order_import.xlsx"
    return path if path.is_file() else None
