"""出货单 Excel 导出：版式对齐打印预览。"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

INK = "111111"
MUTED = "555555"
HEADER_BG = "F3F4F6"
LINE = "333333"
DRAFT = "C45656"
VOID = "888888"
COLS = 5

thin_side = Side(style="thin", color=LINE)
thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)


def _font(bold: bool = False, size: int = 11, color: str = INK) -> Font:
    return Font(name="微软雅黑", bold=bold, size=size, color=color)


def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _set_widths(ws: Worksheet, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _txt(v: Any, default: str = "—") -> str:
    if v is None:
        return default
    s = str(v).strip()
    return s if s else default


def _date_txt(v: Any) -> str:
    if v is None:
        return "—"
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).replace("T", " ")
    return s[:10] if s else "—"


def _num(v: Any) -> float | None:
    if v is None:
        return None
    try:
        return float(Decimal(str(v)))
    except Exception:
        return None


def _merge_row(ws: Worksheet, row: int, text: str, *, bold: bool = False, size: int = 11, color: str = INK) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
    cell = ws.cell(row, 1, text)
    cell.font = _font(bold, size, color)
    cell.alignment = _align("left", wrap=True)


def _clear_header_footer(ws: Worksheet) -> None:
    """去掉 Excel/WPS 页眉页脚（含页码、日期等默认项）。"""
    for block in (
        ws.oddHeader,
        ws.oddFooter,
        ws.evenHeader,
        ws.evenFooter,
        ws.firstHeader,
        ws.firstFooter,
    ):
        block.left.text = ""
        block.center.text = ""
        block.right.text = ""


def _apply_print_setup(ws: Worksheet, last_row: int) -> None:
    """A4 纵向；列宽按纸面设计，避免依赖「放大适配」（WPS/Excel 常只缩小不放大）。"""
    ws.print_area = f"A1:{get_column_letter(COLS)}{last_row}"
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    # 不用 fitToPage 放大：列宽已按 A4 铺满；仅在内容超宽时由用户缩放
    ws.page_setup.fitToPage = False
    ws.page_setup.scale = 100
    ws.sheet_properties.pageSetUpPr.fitToPage = False
    ws.page_margins.left = 0.45
    ws.page_margins.right = 0.45
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    _clear_header_footer(ws)


def _write_shipment_sheet(ws: Worksheet, detail: dict) -> None:
    ws.sheet_view.showGridLines = False
    # A4+窄边距下约 100～110 字符宽较满；过宽会被裁切或被迫缩小
    _set_widths(ws, {1: 12, 2: 28, 3: 22, 4: 20, 5: 26})

    status = str(detail.get("status") or "")
    row = 1
    if status == "draft":
        _merge_row(ws, row, "【草稿】未确认出货，仅供内部参考", bold=True, size=12, color=DRAFT)
        ws.cell(row, 1).alignment = _align("center")
        row += 1
    elif status == "void":
        _merge_row(ws, row, "【已作废】本单作废，不作发货凭证", bold=True, size=12, color=VOID)
        ws.cell(row, 1).alignment = _align("center")
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
    title = ws.cell(row, 1, "出 货 单")
    title.font = _font(True, 20)
    title.alignment = _align("center")
    ws.row_dimensions[row].height = 32
    row += 2

    # 发货方
    _merge_row(ws, row, _txt(detail.get("seller_name")), bold=True, size=14)
    row += 1
    _merge_row(
        ws,
        row,
        f"联系人：{_txt(detail.get('seller_contact_person'))}　　"
        f"电话：{_txt(detail.get('seller_contact_mobile'))}",
        size=10,
    )
    row += 1
    _merge_row(ws, row, f"地址：{_txt(detail.get('seller_address'))}", size=10)
    row += 2

    meta = [
        (f"出货单号：{_txt(detail.get('shipment_no'))}", f"出货日期：{_date_txt(detail.get('ship_date'))}"),
        (f"销售单：{_txt(detail.get('sales_order_no'))}", f"生产单：{_txt(detail.get('header_no') or detail.get('order_no'))}"),
        (f"货号：{_txt(detail.get('product_code'))}", ""),
        (
            f"物流：{_txt(detail.get('logistics_company'))}",
            f"运单号：{_txt(detail.get('tracking_no'))}",
        ),
    ]
    for left, right in meta:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row, 1, left)
        c.font = _font(False, 11)
        c.alignment = _align("left")
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=COLS)
        c2 = ws.cell(row, 4, right)
        c2.font = _font(False, 11)
        c2.alignment = _align("left")
        ws.row_dimensions[row].height = 20
        row += 1

    row += 1
    _merge_row(ws, row, f"收货单位：{_txt(detail.get('customer_name'))}", bold=True, size=11)
    row += 1
    contact_mobile = _txt(detail.get("customer_contact_mobile"), "")
    contact_line = f"联系人：{_txt(detail.get('customer_contact_name'))}"
    if contact_mobile and contact_mobile != "—":
        contact_line = f"{contact_line}　　电话：{contact_mobile}"
    _merge_row(ws, row, contact_line, size=11)
    row += 1
    _merge_row(ws, row, f"收货地址：{_txt(detail.get('customer_address'))}", size=11)
    row += 2

    headers = ["序号", "货号", "颜色", "尺码", "数量"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row, col, h)
        cell.font = _font(True, 11)
        cell.fill = _fill(HEADER_BG)
        cell.border = thin
        if col in (1, 3, 4):
            cell.alignment = _align("center")
        elif col == 5:
            cell.alignment = _align("right")
        else:
            cell.alignment = _align("left")
    header_row = row
    ws.row_dimensions[row].height = 22
    row += 1

    for ln in detail.get("lines") or []:
        vals = [
            ln.get("seq") or "",
            _txt(ln.get("product_code") or detail.get("product_code"), ""),
            _txt(ln.get("color_name")),
            _txt(ln.get("size_value")),
            _num(ln.get("qty")),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row, col, val if val is not None else "—")
            cell.font = _font(False, 11)
            cell.border = thin
            if col in (1, 3, 4):
                cell.alignment = _align("center")
            elif col == 5:
                cell.alignment = _align("right")
                if isinstance(val, (int, float)):
                    cell.number_format = "0"
            else:
                cell.alignment = _align("left")
        ws.row_dimensions[row].height = 20
        row += 1

    if not (detail.get("lines") or []):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
        cell = ws.cell(row, 1, "（无明细）")
        cell.border = thin
        cell.alignment = _align("center")
        row += 1

    row += 1
    total_qty = int(_num(detail.get("total_qty")) or 0)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
    cell = ws.cell(row, 1, f"合计数量：{total_qty} 双")
    cell.font = _font(True, 12)
    cell.alignment = _align("right")
    row += 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=COLS)
    cell = ws.cell(row, 1, f"备注：{_txt(detail.get('notes'), '无')}")
    cell.font = _font(False, 10, MUTED)
    cell.alignment = _align("left", "top", wrap=True)
    ws.row_dimensions[row].height = 18
    ws.row_dimensions[row + 1].height = 18
    row += 3

    # 签字区：左右各半
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row, 1, "发货方签字/盖章").font = _font(False, 10, MUTED)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=COLS)
    ws.cell(row, 4, "收货方签字/盖章").font = _font(False, 10, MUTED)
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row, 1, "________________").font = _font(False, 11)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=COLS)
    ws.cell(row, 4, "________________").font = _font(False, 11)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row, 1, "日期：________").font = _font(False, 10, MUTED)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=COLS)
    ws.cell(row, 4, "日期：________").font = _font(False, 10, MUTED)

    ws.print_title_rows = f"{header_row}:{header_row}"
    _apply_print_setup(ws, row)


def build_shipment_workbook(detail: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "出货单"
    _write_shipment_sheet(ws, detail)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
