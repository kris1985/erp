"""产品开发 Excel 导出：单产品成本明细、批量报价单。"""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image as PILImage

from app.config import get_settings

# 与后台主色一致
ACCENT = "0076FF"
ACCENT_DARK = "005ECC"
INK = "111827"
MUTED = "64748B"
LINE = "CBD5E1"
SOFT = "EFF6FF"
SOFT_ALT = "F8FAFC"
WHITE = "FFFFFF"
TOTAL_BG = "DBEAFE"
SECTION_BG = "F1F5F9"
CARD_BG = "F8FAFC"

thin_side = Side(style="thin", color=LINE)
med_side = Side(style="medium", color=ACCENT)
thick_side = Side(style="medium", color=ACCENT_DARK)

thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
money_fmt = "#,##0.00"
qty_fmt = "0.00"

COLS = 9  # A-I


def _d(v: Any) -> Decimal:
    try:
        return Decimal(str(v or 0))
    except Exception:
        return Decimal("0")


def _f(v: Any) -> float:
    return float(_d(v))


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, size: int = 11, color: str = INK, name: str = "微软雅黑") -> Font:
    return Font(name=name, bold=bold, size=size, color=color)


def _set_widths(ws: Worksheet, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _resolve_upload(url: str | None) -> Path | None:
    if not url:
        return None
    raw = str(url).strip()
    if not raw:
        return None
    settings = get_settings()
    uploads = Path(settings.uploads_dir)
    if raw.startswith("/uploads/"):
        path = uploads / Path(raw).name
        return path if path.is_file() else None
    if raw.startswith("uploads/"):
        path = uploads / Path(raw).name
        return path if path.is_file() else None
    p = Path(raw)
    if p.is_file():
        return p
    cand = uploads / Path(raw).name
    return cand if cand.is_file() else None


def _make_thumb_bytes(path: Path, max_w: int, max_h: int) -> BytesIO | None:
    try:
        with PILImage.open(path) as im:
            im = im.convert("RGB")
            im.thumbnail((max_w, max_h), PILImage.Resampling.LANCZOS)
            buf = BytesIO()
            im.save(buf, format="JPEG", quality=88)
            buf.seek(0)
            return buf
    except Exception:
        return None


def _add_image(ws: Worksheet, path: Path | None, anchor: str, max_w: int, max_h: int) -> bool:
    if not path:
        return False
    buf = _make_thumb_bytes(path, max_w, max_h)
    if not buf:
        return False
    try:
        xl = XLImage(buf)
        xl.anchor = anchor
        ws.add_image(xl)
        return True
    except Exception:
        return False


def _col_width_px(ws: Worksheet, col: int) -> float:
    letter = get_column_letter(col)
    width = ws.column_dimensions[letter].width
    if width is None:
        width = 8.43
    # Excel 列宽字符 → 像素近似
    return float(width) * 7.0 + 5.0


def _row_height_px(ws: Worksheet, row: int) -> float:
    height = ws.row_dimensions[row].height
    if height is None:
        height = 15.0
    return float(height) * 96.0 / 72.0


def _add_cell_image(
    ws: Worksheet,
    path: Path | None,
    row: int,
    col: int,
    max_w: int,
    max_h: int,
) -> bool:
    """将缩略图放入指定单元格并水平/垂直居中。"""
    return _add_range_image(ws, path, row, col, row, col, max_w, max_h)


def _add_range_image(
    ws: Worksheet,
    path: Path | None,
    row1: int,
    col1: int,
    row2: int,
    col2: int,
    max_w: int,
    max_h: int,
) -> bool:
    """将图片放入合并区域（row1:col1 ~ row2:col2）并居中。"""
    if not path:
        return False
    buf = _make_thumb_bytes(path, max_w, max_h)
    if not buf:
        return False
    try:
        xl = XLImage(buf)
        img_w = int(xl.width or max_w)
        img_h = int(xl.height or max_h)
        area_w = sum(_col_width_px(ws, c) for c in range(col1, col2 + 1))
        area_h = sum(_row_height_px(ws, r) for r in range(row1, row2 + 1))
        pad = 6.0
        avail_w = max(1.0, area_w - pad * 2)
        avail_h = max(1.0, area_h - pad * 2)
        scale = min(1.0, avail_w / img_w, avail_h / img_h)
        disp_w = max(1, int(img_w * scale))
        disp_h = max(1, int(img_h * scale))
        xl.width = disp_w
        xl.height = disp_h
        col_off = pixels_to_EMU(max(0.0, (area_w - disp_w) / 2))
        row_off = pixels_to_EMU(max(0.0, (area_h - disp_h) / 2))
        marker = AnchorMarker(col=col1 - 1, colOff=int(col_off), row=row1 - 1, rowOff=int(row_off))
        xl.anchor = OneCellAnchor(
            _from=marker,
            ext=XDRPositiveSize2D(cx=int(pixels_to_EMU(disp_w)), cy=int(pixels_to_EMU(disp_h))),
        )
        ws.add_image(xl)
        return True
    except Exception:
        return False


def _paint_range(ws: Worksheet, r1: int, c1: int, r2: int, c2: int, fill: PatternFill | None = None) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            if fill:
                cell.fill = fill
            if not cell.border or cell.border == Border():
                cell.border = thin


def _header_row(ws: Worksheet, row: int, headers: list[str], start_col: int = 1) -> None:
    fill = _fill(ACCENT)
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=h)
        cell.font = _font(bold=True, size=10, color=WHITE)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[row].height = 26


def _section_bar(ws: Worksheet, row: int, title: str, cols: int = COLS) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=f"  {title}")
    cell.font = _font(bold=True, size=12, color=ACCENT_DARK)
    cell.fill = _fill(SECTION_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = thin
    for c in range(2, cols + 1):
        x = ws.cell(row=row, column=c)
        x.fill = _fill(SECTION_BG)
        x.border = thin
    ws.row_dimensions[row].height = 26
    return row + 1


def _apply_outer_border(ws: Worksheet, r1: int, c1: int, r2: int, c2: int) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cur = cell.border or Border()
            cell.border = Border(
                left=med_side if c == c1 else (cur.left or thin_side),
                right=med_side if c == c2 else (cur.right or thin_side),
                top=med_side if r == r1 else (cur.top or thin_side),
                bottom=med_side if r == r2 else (cur.bottom or thin_side),
            )


def build_own_product_workbook(product: dict, partner_id: int | None = None) -> bytes:
    """导出单个产品成本明细（不含报价）。partner_id 保留兼容，已忽略。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "产品成本明细"
    end_row = _fill_product_sheet(ws, product)
    _apply_outer_border(ws, 1, 1, end_row, COLS)
    ws.print_area = f"A1:{get_column_letter(COLS)}{end_row}"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def resolve_batch_quote_price(product: dict, partner_id: int | None) -> float | None:
    """有客户报价用客户价，否则统一报价。"""
    if partner_id is not None:
        for q in product.get("quotes") or []:
            if int(q.get("partner_id") or 0) == int(partner_id):
                return _f(q.get("quote_price"))
    unified = product.get("quote_price")
    if unified is None or unified == "":
        return None
    return _f(unified)


def build_batch_quote_workbook(
    products: list[dict],
    partner_id: int | None = None,
    customer_name: str | None = None,
    company_name: str | None = None,
) -> bytes:
    """批量报价单：图片、编号、颜色、价格。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "产品报价单"
    _set_widths(ws, {1: 8, 2: 16, 3: 24, 4: 28, 5: 14})

    # 标题居中
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    t = ws.cell(row=1, column=1, value="产品报价单")
    t.font = _font(bold=True, size=16, color=ACCENT_DARK)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # 客户名称与表格左对齐，字号更大
    customer_label = f"客户：{customer_name}" if customer_name else "统一报价"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    c = ws.cell(row=2, column=1, value=customer_label)
    c.font = _font(bold=True, size=14, color=INK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 26

    headers = ["序号", "图片", "编号", "颜色", "价格"]
    header_aligns = ["center", "center", "left", "left", "right"]
    header_row = 4
    for col, (h, align) in enumerate(zip(headers, header_aligns), 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = _font(bold=True, size=11, color=WHITE)
        cell.fill = _fill(ACCENT)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = thin
    ws.row_dimensions[header_row].height = 24

    row = header_row + 1
    for i, p in enumerate(products, 1):
        colors = "、".join(
            c.get("name") or "" for c in (p.get("colors") or []) if c.get("name")
        ) or "—"
        price = resolve_batch_quote_price(p, partner_id)

        c1 = ws.cell(row=row, column=1, value=i)
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.border = thin

        c2 = ws.cell(row=row, column=2, value="")
        c2.border = thin
        c2.alignment = Alignment(horizontal="center", vertical="center")

        c3 = ws.cell(row=row, column=3, value=p.get("product_code") or "—")
        c3.font = _font(bold=True, size=11)
        c3.alignment = Alignment(horizontal="left", vertical="center")
        c3.border = thin

        c4 = ws.cell(row=row, column=4, value=colors)
        c4.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c4.border = thin

        c5 = ws.cell(row=row, column=5, value=price if price is not None else "—")
        if price is not None:
            c5.number_format = money_fmt
        c5.font = _font(bold=True, size=12, color=ACCENT_DARK)
        c5.alignment = Alignment(horizontal="right", vertical="center")
        c5.border = thin

        if i % 2 == 0:
            alt = _fill(SOFT_ALT)
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = alt

        ws.row_dimensions[row].height = 58
        path = _resolve_upload(p.get("image_url"))
        if path:
            thumb = _make_thumb_bytes(path, 72, 72)
            if thumb:
                try:
                    img = XLImage(thumb)
                    img.width = 52
                    img.height = 52
                    img.anchor = f"B{row}"
                    ws.add_image(img)
                except Exception:
                    pass
        row += 1

    end_row = row - 1
    _apply_outer_border(ws, header_row, 1, end_row, 5)

    # 底部：共 n 款（居中）+ 落款公司/日期（右下）
    count_row = end_row + 2
    ws.merge_cells(start_row=count_row, start_column=1, end_row=count_row, end_column=5)
    count_cell = ws.cell(row=count_row, column=1, value=f"共 {len(products)} 款")
    count_cell.font = _font(size=11, color=MUTED)
    count_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[count_row].height = 22

    sign_date = datetime.now().strftime("%Y-%m-%d")
    company = (company_name or "").strip() or "—"
    sign_row = count_row + 2
    ws.merge_cells(start_row=sign_row, start_column=3, end_row=sign_row, end_column=5)
    company_cell = ws.cell(row=sign_row, column=3, value=company)
    company_cell.font = _font(bold=True, size=12, color=INK)
    company_cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[sign_row].height = 22

    date_row = sign_row + 1
    ws.merge_cells(start_row=date_row, start_column=3, end_row=date_row, end_column=5)
    date_cell = ws.cell(row=date_row, column=3, value=sign_date)
    date_cell.font = _font(size=11, color=MUTED)
    date_cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[date_row].height = 20

    last_row = date_row
    ws.print_area = f"A1:E{last_row}"
    # 打印预览水平居中（内容块相对纸张）
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    # openpyxl 需同时打开 sheet 级 fitToPage，否则 Excel 可能忽略缩放
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.6
    ws.page_margins.right = 0.6
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.6
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fill_product_sheet(ws: Worksheet, p: dict, partner_id: int | None = None) -> int:
    code = p.get("product_code") or "—"
    colors = "、".join(c.get("name") or "" for c in (p.get("colors") or []) if c.get("name")) or "—"
    mat = _d(p.get("material_cost"))
    lab = _d(p.get("labor_cost"))
    oth = _d(p.get("other_cost"))
    tot = mat + lab + oth

    _set_widths(
        ws,
        {
            1: 12,
            2: 16,
            3: 10,
            4: 14,
            5: 10,
            6: 9,
            7: 9,
            8: 14,
            9: 11,
        },
    )

    # —— 标题条 ——
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COLS)
    title = ws.cell(row=1, column=1, value="产品成本明细")
    title.font = _font(bold=True, size=20, color=WHITE)
    title.fill = _fill(ACCENT)
    title.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(2, COLS + 1):
        ws.cell(row=1, column=c).fill = _fill(ACCENT)
    ws.row_dimensions[1].height = 40

    # —— 产品头：左图（4 行）右信息（4 行等高）——
    info_rows = [
        ("产品编号", code),
        ("颜色", colors),
        ("材料 / 人工 / 其它", f"{_f(mat):.2f}  /  {_f(lab):.2f}  /  {_f(oth):.2f}"),
        ("总成本", _f(tot)),
    ]
    info_start = 2
    info_end = info_start + len(info_rows) - 1  # 正好 4 行
    for r in range(info_start, info_end + 1):
        ws.row_dimensions[r].height = 36
        for c in range(1, COLS + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = _fill(CARD_BG)
            cell.border = thin

    # 产品图：合并 A:C × 4 行，与右侧文本等高并居中
    ws.merge_cells(start_row=info_start, start_column=1, end_row=info_end, end_column=3)
    img_cell = ws.cell(row=info_start, column=1, value="")
    img_cell.alignment = Alignment(horizontal="center", vertical="center")
    img_path = _resolve_upload(p.get("image_url"))
    has_img = _add_range_image(ws, img_path, info_start, 1, info_end, 3, 220, 200)
    if not has_img:
        img_cell.value = "暂无产品图"
        img_cell.font = _font(size=11, color=MUTED)

    for i, (label, val) in enumerate(info_rows):
        r = info_start + i
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        lab_cell = ws.cell(row=r, column=4, value=label)
        lab_cell.font = _font(bold=True, size=10, color=MUTED)
        lab_cell.fill = _fill(SOFT)
        lab_cell.alignment = Alignment(horizontal="center", vertical="center")
        lab_cell.border = thin
        ws.cell(row=r, column=5).fill = _fill(SOFT)
        ws.cell(row=r, column=5).border = thin

        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=COLS)
        val_cell = ws.cell(row=r, column=6, value=val)
        is_total = label == "总成本"
        val_bg = _fill(TOTAL_BG if is_total else WHITE)
        val_cell.font = _font(
            bold=is_total,
            size=13 if is_total else 11,
            color=ACCENT_DARK if is_total else INK,
        )
        val_cell.fill = val_bg
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        val_cell.border = thin
        if isinstance(val, float):
            val_cell.number_format = money_fmt
        for c in range(7, COLS + 1):
            x = ws.cell(row=r, column=c)
            x.fill = val_bg
            x.border = thin
            x.alignment = Alignment(horizontal="center", vertical="center")

    row = info_end + 1

    # —— 物料明细（含缩略图）——
    row = _section_bar(ws, row, "物料明细")
    headers = ["图片", "名称", "颜色", "单价", "数量", "单位", "材料总价"]
    _header_row(ws, row, headers)
    for c in range(len(headers) + 1, COLS + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill(ACCENT)
        cell.border = thin
    row += 1

    materials = list(p.get("materials") or [])
    if not materials:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
        empty = ws.cell(row=row, column=1, value="暂无物料")
        empty.font = _font(size=10, color=MUTED)
        empty.alignment = Alignment(horizontal="center", vertical="center")
        _paint_range(ws, row, 1, row, COLS, _fill(WHITE))
        ws.row_dimensions[row].height = 28
        row += 1
    else:
        for mi, m in enumerate(materials, 1):
            bg = _fill(SOFT_ALT if mi % 2 == 0 else WHITE)
            ws.row_dimensions[row].height = 48
            vals = [
                "",
                m.get("supplier_product_name") or "—",
                m.get("color_name") or "—",
                _f(m.get("unit_price")),
                _f(m.get("qty")),
                m.get("pricing_unit_name") or "—",
                _f(m.get("line_total")),
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.border = thin
                cell.fill = bg
                cell.font = _font(size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if c in (4, 5, 7):
                    cell.number_format = money_fmt if c != 5 else qty_fmt
                if c == 7:
                    cell.font = _font(bold=True, size=10, color=ACCENT_DARK)
            for c in range(len(vals) + 1, COLS + 1):
                cell = ws.cell(row=row, column=c)
                cell.border = thin
                cell.fill = bg
            m_path = _resolve_upload(m.get("image_url"))
            if not _add_cell_image(ws, m_path, row, 1, 40, 40):
                ws.cell(row=row, column=1).value = "—"
                ws.cell(row=row, column=1).font = _font(size=10, color=MUTED)
            row += 1

        for c in range(1, 6):
            cell = ws.cell(row=row, column=c)
            cell.fill = _fill(TOTAL_BG)
            cell.border = thin
        lab_cell = ws.cell(row=row, column=6, value="材料小计")
        lab_cell.font = _font(bold=True, size=10, color=MUTED)
        lab_cell.fill = _fill(TOTAL_BG)
        lab_cell.border = thin
        lab_cell.alignment = Alignment(horizontal="center", vertical="center")
        tot_cell = ws.cell(row=row, column=7, value=_f(mat))
        tot_cell.font = _font(bold=True, size=11, color=ACCENT_DARK)
        tot_cell.fill = _fill(TOTAL_BG)
        tot_cell.border = thin
        tot_cell.number_format = money_fmt
        tot_cell.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(8, COLS + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = _fill(TOTAL_BG)
            cell.border = thin
        ws.row_dimensions[row].height = 26
        row += 1

    # —— 人工 + 其它（左右两栏）——
    row = _section_bar(ws, row, "人工成本  /  其它成本")
    # sub headers
    for c, h in enumerate(["工序", "价格(元)", "", "项目", "金额(元)", "", "", "", ""], 1):
        cell = ws.cell(row=row, column=c, value=h if h else None)
        if h:
            cell.font = _font(bold=True, size=10, color=WHITE)
            cell.fill = _fill(ACCENT_DARK)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.fill = _fill(SECTION_BG)
        cell.border = thin
    # merge empty spacer cols visually
    ws.cell(row=row, column=3).fill = _fill(SECTION_BG)
    for c in range(6, COLS + 1):
        ws.cell(row=row, column=c).fill = _fill(SECTION_BG)
    ws.row_dimensions[row].height = 24
    row += 1

    labors = list(p.get("labors") or [])
    others = list(p.get("other_costs") or [])
    n = max(len(labors), len(others), 1)
    for i in range(n):
        bg = _fill(SOFT_ALT if (i + 1) % 2 == 0 else WHITE)

        if i < len(labors):
            l = labors[i]
            a = ws.cell(row=row, column=1, value=l.get("process_name") or "—")
            a.font = _font(size=10)
            a.alignment = Alignment(horizontal="center", vertical="center")
            b = ws.cell(row=row, column=2, value=_f(l.get("unit_price")))
            b.font = _font(size=10)
            b.number_format = money_fmt
            b.alignment = Alignment(horizontal="center", vertical="center")
        elif i == 0 and not labors:
            a = ws.cell(row=row, column=1, value="暂无工序")
            a.font = _font(size=10, color=MUTED)
            a.alignment = Alignment(horizontal="center", vertical="center")
            b = ws.cell(row=row, column=2, value="")
            b.font = _font(size=10)
        else:
            a = ws.cell(row=row, column=1, value="")
            a.font = _font(size=10)
            b = ws.cell(row=row, column=2, value="")
            b.font = _font(size=10)
        a.fill = bg
        a.border = thin
        b.fill = bg
        b.border = thin
        b.alignment = Alignment(horizontal="center", vertical="center")

        spacer = ws.cell(row=row, column=3, value="")
        spacer.fill = _fill(SECTION_BG)
        spacer.border = thin

        if i < len(others):
            o = others[i]
            c4 = ws.cell(row=row, column=4, value=o.get("name") or "—")
            c4.font = _font(size=10)
            c4.alignment = Alignment(horizontal="center", vertical="center")
            c5 = ws.cell(row=row, column=5, value=_f(o.get("amount")))
            c5.font = _font(size=10)
            c5.number_format = money_fmt
            c5.alignment = Alignment(horizontal="center", vertical="center")
        elif i == 0 and not others:
            c4 = ws.cell(row=row, column=4, value="暂无其它成本")
            c4.font = _font(size=10, color=MUTED)
            c4.alignment = Alignment(horizontal="center", vertical="center")
            c5 = ws.cell(row=row, column=5, value="")
            c5.font = _font(size=10)
        else:
            c4 = ws.cell(row=row, column=4, value="")
            c4.font = _font(size=10)
            c5 = ws.cell(row=row, column=5, value="")
            c5.font = _font(size=10)
        c4.fill = bg
        c4.border = thin
        c4.alignment = Alignment(horizontal="center", vertical="center")
        c5.fill = bg
        c5.border = thin
        c5.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(6, COLS + 1):
            x = ws.cell(row=row, column=c)
            x.fill = _fill(SECTION_BG)
            x.border = thin
        ws.row_dimensions[row].height = 24
        row += 1

    # subtotals for labor/other
    for c in range(1, COLS + 1):
        ws.cell(row=row, column=c).fill = _fill(TOTAL_BG)
        ws.cell(row=row, column=c).border = thin
        ws.cell(row=row, column=c).alignment = Alignment(horizontal="center", vertical="center")
    lab1 = ws.cell(row=row, column=1, value="人工小计")
    lab1.font = _font(bold=True, size=10, color=MUTED)
    lab1.fill = _fill(TOTAL_BG)
    lab1.border = thin
    lab1.alignment = Alignment(horizontal="center", vertical="center")
    b = ws.cell(row=row, column=2, value=_f(lab))
    b.font = _font(bold=True, size=10, color=ACCENT_DARK)
    b.number_format = money_fmt
    b.fill = _fill(TOTAL_BG)
    b.border = thin
    b.alignment = Alignment(horizontal="center", vertical="center")
    lab2 = ws.cell(row=row, column=4, value="其它小计")
    lab2.font = _font(bold=True, size=10, color=MUTED)
    lab2.fill = _fill(TOTAL_BG)
    lab2.border = thin
    lab2.alignment = Alignment(horizontal="center", vertical="center")
    c5 = ws.cell(row=row, column=5, value=_f(oth))
    c5.font = _font(bold=True, size=10, color=ACCENT_DARK)
    c5.number_format = money_fmt
    c5.fill = _fill(TOTAL_BG)
    c5.border = thin
    c5.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 26

    return row
