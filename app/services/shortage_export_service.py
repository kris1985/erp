"""P1-5：缺料批量导出 Excel + 企微/钉钉 Webhook 催办摘要。"""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Order, OwnProduct
from app.services import im_alerts_service, material_service, order_risk, progress_service
from app.services.purchase_service import annotate_rows_with_etas


HEADERS = [
    "生产单号",
    "款号",
    "客户",
    "交期",
    "急单",
    "风险等级",
    "风险原因",
    "预计齐套日",
    "物料编码",
    "物料名称",
    "尺码",
    "缺口",
    "待购",
    "供应商",
    "预计到料日",
]


def _as_date_str(v: Any) -> str:
    if v is None or v == "":
        return ""
    if hasattr(v, "isoformat"):
        return v.isoformat()[:10]
    return str(v)[:10]


def build_shortage_export_rows(
    db: Session,
    tenant_id: int,
    *,
    order_ids: list[int] | None = None,
    keyword: str | None = None,
    partner_id: int | None = None,
    rush_only: bool = False,
    hide_purchased: bool = True,
    limit: int = 2000,
) -> list[dict[str, Any]]:
    rows = material_service.list_shortages(
        db,
        tenant_id,
        order_ids=order_ids,
        keyword=keyword,
        partner_id=partner_id,
        rush_only=rush_only,
        hide_purchased=hide_purchased,
    )
    if not rows:
        return []
    eta = annotate_rows_with_etas(db, tenant_id, rows)
    kit_by_order = eta.get("by_order_id") or {}

    oids = sorted({int(r["order_id"]) for r in rows if r.get("order_id")})
    orders = {
        o.id: o
        for o in db.scalars(select(Order).where(Order.tenant_id == tenant_id, Order.id.in_(oids))).all()
    }
    product_ids = {o.own_product_id for o in orders.values() if o.own_product_id}
    products = {
        p.id: p
        for p in db.scalars(
            select(OwnProduct).where(OwnProduct.tenant_id == tenant_id, OwnProduct.id.in_(product_ids))
        ).all()
    } if product_ids else {}

    board = progress_service.progress_board(db, tenant_id)
    board_by_id = {int(o["id"]): o for o in (board.get("orders") or []) if o.get("id") is not None}
    kits = material_service.order_kit_summaries(db, tenant_id, oids)

    risk_cache: dict[int, dict[str, Any]] = {}
    for oid in oids:
        o = orders.get(oid)
        if not o:
            continue
        board_o = board_by_id.get(oid) or {}
        kit = kits.get(oid) or {}
        ready = kit_by_order.get(str(oid)) or kit_by_order.get(oid) or kit.get("kit_ready_date")
        risk_cache[oid] = order_risk.compute_order_risk(
            status=o.status,
            delivery_date=o.delivery_date,
            overall_percent=float(board_o.get("overall_percent") or 0),
            is_rush=bool(o.is_rush),
            kit_ok=kit.get("kit_ok"),
            kit_ready_date=_as_date_str(ready) or None,
        )

    out: list[dict[str, Any]] = []
    for r in rows[:limit]:
        oid = int(r["order_id"])
        o = orders.get(oid)
        product = products.get(o.own_product_id) if o and o.own_product_id else None
        risk = risk_cache.get(oid) or {}
        reasons = risk.get("risk_reasons") or []
        ready = kit_by_order.get(str(oid)) or kit_by_order.get(oid)
        out.append(
            {
                "order_id": oid,
                "order_no": r.get("order_no") or (o.order_no if o else None),
                "product_code": product.product_code if product else None,
                "customer_name": o.customer_name if o else None,
                "delivery_date": _as_date_str(o.delivery_date if o else None),
                "is_rush": bool(r.get("is_rush") or (o.is_rush if o else False)),
                "risk_level": risk.get("risk_level") or "none",
                "risk_label": risk.get("risk_label") or "—",
                "risk_reason_text": "；".join(x.get("text") or "" for x in reasons if x.get("text")),
                "kit_ready_date": _as_date_str(ready),
                "supplier_product_code": r.get("supplier_product_code"),
                "supplier_product_name": r.get("supplier_product_name"),
                "size_value": r.get("size_value"),
                "shortage_qty": float(r.get("shortage_qty") or 0),
                "to_buy_qty": float(r.get("to_buy_qty") or 0),
                "partner_name": r.get("partner_name"),
                "expected_ready_date": _as_date_str(r.get("expected_ready_date")),
            }
        )
    return out


def build_shortage_workbook(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "缺料催办"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append(
            [
                r.get("order_no") or "",
                r.get("product_code") or "",
                r.get("customer_name") or "",
                r.get("delivery_date") or "",
                "是" if r.get("is_rush") else "",
                r.get("risk_label") or r.get("risk_level") or "",
                r.get("risk_reason_text") or "",
                r.get("kit_ready_date") or "",
                r.get("supplier_product_code") or "",
                r.get("supplier_product_name") or "",
                r.get("size_value") or "",
                r.get("shortage_qty") or 0,
                r.get("to_buy_qty") or 0,
                r.get("partner_name") or "",
                r.get("expected_ready_date") or "",
            ]
        )
    from openpyxl.utils import get_column_letter

    widths = [14, 14, 12, 12, 6, 10, 28, 12, 14, 18, 8, 10, 10, 14, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_shortage_push_message(
    db: Session,
    tenant_id: int,
    rows: list[dict[str, Any]],
) -> dict[str, Any]:
    from app.models import Tenant

    tenant = db.get(Tenant, tenant_id)
    factory = tenant.name if tenant else "工厂"
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    order_nos = []
    seen: set[str] = set()
    for r in rows:
        no = str(r.get("order_no") or "")
        if no and no not in seen:
            seen.add(no)
            order_nos.append(no)
    red_n = sum(1 for r in rows if r.get("risk_level") == "red")
    lines = [
        f"【{factory}】缺料催办 {now}",
        f"缺料行 {len(rows)}，涉及生产单 {len(order_nos)}，其中高风险行 {red_n}",
    ]
    for r in rows[:8]:
        lines.append(
            f"- {r.get('order_no')} {r.get('product_code') or '—'} "
            f"{r.get('supplier_product_name') or r.get('supplier_product_code') or '物料'} "
            f"缺口{r.get('shortage_qty')} "
            f"齐套日{r.get('kit_ready_date') or '—'} "
            f"[{r.get('risk_label') or r.get('risk_level')}]"
        )
    if len(rows) > 8:
        lines.append(f"…其余 {len(rows) - 8} 行见导出 Excel")
    content = "\n".join(lines)
    return {
        "kind": "shortage_export",
        "generated_at": now,
        "row_count": len(rows),
        "order_count": len(order_nos),
        "message": {"msgtype": "text", "text": {"content": content}},
    }


def push_shortage_digest(
    db: Session,
    tenant_id: int,
    rows: list[dict[str, Any]],
    *,
    webhook_url_override: str | None = None,
) -> dict[str, Any]:
    settings = im_alerts_service.get_im_alerts_by_tenant_id(db, tenant_id)
    url = im_alerts_service._clean_url(webhook_url_override) or settings.get("webhook_url")
    if not url:
        raise ValueError("未配置企微/钉钉 Webhook，请先在「IM 预警」填写地址")
    payload = build_shortage_push_message(db, tenant_id, rows)
    result = im_alerts_service.post_json(url, payload["message"])
    return {"payload": payload, "result": result, "webhook_url": url}
